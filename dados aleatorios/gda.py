import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import pdfplumber
import os

# Função para ler os dados do arquivo .pdf
def ler_dados_pdf(arquivo):
    dados = {
        "Pessoa": [],
        "Idade": [],
        "Peso": [],
        "Dinheiro": []
    }

    with pdfplumber.open(arquivo) as pdf:
        for page in pdf.pages:
            linhas = page.extract_text().splitlines()
            for linha in linhas:
                linha = linha.strip()
                if linha:  # Ignorar linhas vazias
                    partes = linha.split(",")  # Dividir os valores por vírgula
                    if len(partes) >= 4:  # Certificar-se de que a linha tem todas as colunas esperadas
                        dados["Pessoa"].append(partes[0].strip())
                        dados["Idade"].append(int(partes[1].strip()))
                        dados["Peso"].append(partes[2].strip())
                        # Manter o valor completo do dinheiro exatamente como está no documento
                        valor_dinheiro = ",".join(partes[3:]).strip()
                        dados["Dinheiro"].append(valor_dinheiro)

    return dados

# Definir o caminho do arquivo PDF na mesma pasta do script
pdf_path = os.path.join(os.path.dirname(__file__), "dados.pdf")

# Ler dados do arquivo .pdf
dados = ler_dados_pdf(pdf_path)

# Converter os dados em DataFrame
df = pd.DataFrame(dados)

# Converter Dinheiro para valores numéricos
def convert_dinheiro(valor):
    valor = valor.replace("R$", "").replace(".", "").replace(",", ".")
    if valor.count(".") > 1:  # Corrigir valores com separadores incorretos
        partes = valor.split(".")
        valor = "".join(partes[:-1]) + "." + partes[-1]
    return float(valor)

dinheiro_numerico = df["Dinheiro"].apply(convert_dinheiro)

# Calcular médias
media_idade = df["Idade"].mean()
media_peso = df["Peso"].str.replace("kg", "").astype(int).mean()
media_dinheiro = dinheiro_numerico.mean()

# Adicionar as colunas de médias (MIdade, MPeso, MValor)
df["MIdade"] = ""
df["MPeso"] = ""
df["MValor"] = ""

# Preencher as médias na primeira linha
df.loc[0, "MIdade"] = round(media_idade, 2)
df.loc[0, "MPeso"] = f"{round(media_peso, 2)}kg"
df.loc[0, "MValor"] = f"R${media_dinheiro:,.2f}".replace(",", ".")

# Gerar o arquivo Excel
arquivo_excel = "dados_pessoas.xlsx"
df.to_excel(arquivo_excel, index=False, sheet_name="Dados")

# Ajustar colunas da planilha gerada
wb = openpyxl.load_workbook(arquivo_excel)
sheet = wb["Dados"]

# Preencher cor vermelha para condições específicas
fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
    dinheiro = convert_dinheiro(row[3].value)
    if dinheiro < 50000:
        for cell in row:
            cell.fill = fill_red

# Ajustar largura das colunas e centralizar o conteúdo
for col in sheet.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:  # Calcular o comprimento máximo do conteúdo de cada célula
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2
    sheet.column_dimensions[col_letter].width = adjusted_width

# Centralizar o conteúdo das células
for row in sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save("dados_pessoas.xlsx")
print(f"Planilha 'dados_pessoas.xlsx' gerada e ajustada com sucesso.")
